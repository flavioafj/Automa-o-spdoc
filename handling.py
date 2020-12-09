from openpyxl import load_workbook
import datetime
import sys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


#abrir o e-mail
driver = webdriver.Chrome(executable_path=r'chromedriver.exe')
driver.get("https://webmail.fazenda.sp.gov.br/owa/")
driver.implicitly_wait(3)
driver.find_element(By.CSS_SELECTOR, '[autoid="_n_j"]').click()

#abrir uma nova tab para o SPDOC
windows_before = driver.current_window_handle


driver.execute_script('window.open("https://www.documentos.spsempapel.sp.gov.br/siga/public/app/login?cont=https://www.documentos.spsempapel.sp.gov.br/siga/app/principal")')
WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
windows_after = driver.window_handles
new_window = [x for x in windows_after if x != windows_before][0]
driver.switch_to.window(new_window)
WebDriverWait(driver, 20).until(EC.title_contains("SP"))


#=== SPDOC Autenticação
driver.find_element(By.ID, "username").send_keys("SFP25402")

#senha = input("Enter senha:")
senha = "Bruno1980%"
driver.find_element(By.ID, "password").send_keys(senha + Keys.ENTER)

driver.implicitly_wait(2)
driver.find_element(By.ID, "collapse-header-7").click()

#=== SPDOC escolha do processo na mesa
driver.implicitly_wait(6)
WebDriverWait(driver, timeout=6).until(lambda g: g.find_element(By.CSS_SELECTOR, '#collapsetab-7 > div > div.col-6.col-md-3.mb-2 > div > button.btn.btn-primary.btn-sm'))

driver.find_element(By.CSS_SELECTOR, "#configMenu > button").click()

driver.implicitly_wait(6)

driver.find_element(By.CSS_SELECTOR, "#configMenu > div > div.form-group.pb-2.mb-1.border-bottom > select").send_keys("500" + Keys.ENTER)

driver.implicitly_wait(2)
driver.find_element(By.CSS_SELECTOR, "#collapsetab-7 > div > div.col-6.col-md-3.mb-2 > div > button.btn.btn-primary.btn-sm").click()

#driver.find_element_by_link_text("SFP-EXP-2020/210311-A").click()

driver.switch_to.window(windows_before)
WebDriverWait(driver, 20).until(EC.title_contains("F"))


#Abre-se a planilha
wb = load_workbook(filename = r'C:\\Users\\Cliente\\OneDrive\\Documentos\\SEFAZ\\NFP\\Gerador_de_Oficio2.xlsm', data_only=True, keep_vba=True)

def conta_linha (planilha):
    sheet_ranges = wb[planilha]
    row_count = sheet_ranges.max_row
    conta_linhas = int(0)
    col = int(1)
    if planilha == 'Juiz_Vara_End':
        col = 4
    
    for row in range(1, row_count + 1):
        if sheet_ranges.cell(row, col).value != "" and sheet_ranges.cell(row, col).value != None:
            conta_linhas += 1
    return conta_linhas    



def lista_de_pesquisados(processo, pedido):
    sheet_ranges = wb['Prateleira']
    limite = conta_linha('Prateleira')
    lista_de = list()
    lista_de_nao_transferidos =  list()

    if pedido == 'Transferência':
        for row in range(2, limite + 1):
            
            if sheet_ranges.cell(row, 1).value == processo:
                #serão transferidos 
                if sheet_ranges.cell(row, 3).value > 25 and sheet_ranges.cell(row, 4).value.strip() == 'NÃO INSCRITO NO CADIN':
                    lista_de.append(sheet_ranges.cell(row, 5).value)
                    lista_de.append(sheet_ranges.cell(row, 2).value)
                    lista_de.append(sheet_ranges.cell(row, 3).value)
                    lista_de.append(sheet_ranges.cell(row, 4).value)
                #Não serão transferidos
                else:
                    lista_de_nao_transferidos.append(sheet_ranges.cell(row, 5).value)
                    lista_de_nao_transferidos.append(sheet_ranges.cell(row, 2).value)
                    lista_de_nao_transferidos.append(sheet_ranges.cell(row, 3).value)
                    lista_de_nao_transferidos.append(sheet_ranges.cell(row, 4).value)


        lista_final = str()
        #todos serão transferidos
        if len(lista_de_nao_transferidos) == 0 and len(lista_de) > 0:
            tamanho_da_lista = len(lista_de)
            
            for x in range(0, tamanho_da_lista, 4):
                lista_final = lista_final + "@24@ <tr>"
                lista_final = lista_final + "<td style='text-align:center'>" + lista_de[x] + "</td>"
                lista_final = lista_final + "<td style='text-align:center'>" + lista_de[x + 1] + "</td>"
                lista_final = lista_final + "<td style='text-align:center'>R$ " + str(lista_de[x + 2]) + "</td>"
                lista_final = lista_final + "<td style='text-align:center'>" + lista_de[x + 3] + "</td>"
                lista_final = lista_final + "</tr>"
        
        #Transferência mista
        elif len(lista_de_nao_transferidos) > 0 and len(lista_de) > 0:

            tamanho_da_lista = len(lista_de)
            
            for x in range(0, tamanho_da_lista, 4):
                lista_final = lista_final + "<tr>"
                lista_final = lista_final + "<td style='text-align:center'>" + lista_de[x] + "</td>"
                lista_final = lista_final + "<td style='text-align:center'>" + lista_de[x + 1] + "</td>"
                lista_final = lista_final + "<td style='text-align:center'>R$ " + str(lista_de[x + 2]) + "</td>"
                lista_final = lista_final + "<td style='text-align:center'>" + lista_de[x + 3] + "</td>"
                lista_final = lista_final + "</tr>"

            tamanho_da_lista2 = len(lista_de_nao_transferidos)
            lista_final = lista_final + "</table><br><div><b>Informamos que TODAS a(s) conta(s) já foi(foram) bloqueada(s). No presente momento, por imposição legal, a solicitação de transferência de valores feita por VSª poderá ser atendida SOMENTE COM RELAÇÃO AOS NOMES ACIMA, uma vez que há saldo(s) em conta NFP que é(são) inferior(es) a R$ 25,00 ou há pendências perante o CADIN estadual, relativo(s) a(os) seguinte(s) nome(s):</b></div><br><table><tr><th>NOME</th><th>CPF/CNPJ</th><th>SALDO NO PROGRAMA NFP</th><th>OBSERVAÇÃO</th></tr>"
            
            for y in range(0, tamanho_da_lista2, 4):
                lista_final = lista_final + "<tr>"
                lista_final = lista_final + "<td style='text-align:center'>" + lista_de_nao_transferidos[y] + "</td>"
                lista_final = lista_final + "<td style='text-align:center'>" + lista_de_nao_transferidos[y + 1] + "</td>"
                lista_final = lista_final + "<td style='text-align:center'>R$ " + str(lista_de_nao_transferidos[y + 2]) + "</td>"
                lista_final = lista_final + "<td style='text-align:center'>" + lista_de_nao_transferidos[y + 3] + "</td>"
                lista_final = lista_final + "</tr>"

        #Nenhum nome será transferido
        elif len(lista_de_nao_transferidos) > 0 and len(lista_de) == 0:

            tamanho_da_lista2 = len(lista_de_nao_transferidos)
             
            for y in range(0, tamanho_da_lista2, 4):
                lista_final = lista_final + "@25@ <tr>"
                lista_final = lista_final + "<td style='text-align:center'>" + lista_de_nao_transferidos[y] + "</td>"
                lista_final = lista_final + "<td style='text-align:center'>" + lista_de_nao_transferidos[y + 1] + "</td>"
                lista_final = lista_final + "<td style='text-align:center'>R$ " + str(lista_de_nao_transferidos[y + 2]) + "</td>"
                lista_final = lista_final + "<td style='text-align:center'>" + lista_de_nao_transferidos[y + 3] + "</td>"
                lista_final = lista_final + "</tr>"

    #para os demais pedidos que não são transferências
    else:

        for row in range(1, limite + 1):
            if sheet_ranges.cell(row, 1).value == processo:
                lista_de.append(sheet_ranges.cell(row, 5).value)
                lista_de.append(sheet_ranges.cell(row, 2).value)
                lista_de.append(sheet_ranges.cell(row, 3).value)
                if sheet_ranges.cell(row, 4).value == None :
                    lista_de.append("-")
                else:
                    lista_de.append(sheet_ranges.cell(row, 4).value)

        tamanho_da_lista = len(lista_de)
        lista_final = str()
        for x in range(0, tamanho_da_lista, 4):
            lista_final = lista_final + "<tr>"
            lista_final = lista_final + "<td style='text-align:center'>" + lista_de[x] + "</td>"
            lista_final = lista_final + "<td style='text-align:center'>" + lista_de[x + 1] + "</td>"
            lista_final = lista_final + "<td style='text-align:center'>R$ " + str(lista_de[x + 2]) + "</td>"
            lista_final = lista_final + "<td style='text-align:center'>" + lista_de[x + 3] + "</td>"
            lista_final = lista_final + "</tr>"

    return lista_final

def endereco_juiz(nome, vara):
    sheet_ranges = wb['Juiz_Vara_End']
    limite = conta_linha('Juiz_Vara_End')
    lista_de = list()
    for row in range(1, limite + 1):
        if sheet_ranges.cell(row, 4).value.strip() == nome.strip() and sheet_ranges.cell(row, 5).value.strip() == vara.strip():
            lista_de.append(sheet_ranges.cell(row, 4).value)
            lista_de.append(sheet_ranges.cell(row, 5).value)
            lista_de.append(sheet_ranges.cell(row, 6).value)
            lista_de.append(sheet_ranges.cell(row, 7).value)
            lista_de.append(sheet_ranges.cell(row, 8).value)
     
    return lista_de

def mes():
    x = datetime.datetime.now()
    mes2 = int(x.month)


    mes3 = str()
    if mes2 == 1:
        mes3 = 'janeiro'
        
    
    if mes2 == 2:
        mes3 = 'fevereiro'
        

    if mes2 == 3:
        mes3 = 'março'
        

    if mes2 == 4:
        mes3 = 'abril'
        

    if mes2 == 5:
        mes3 = 'maio'
        

    if mes2 == 6:
        mes3 = 'junho'
        

    if mes2 == 7:
        mes3 = 'julho'
        
    
    if mes2 == 8:
        mes3 = 'agosto'
        

    if mes2 == 9:
        mes3 = 'setembro'
        

    if mes2 ==10:
        mes3 = 'outubro'
        

    if mes2 == 11:
        mes3 = 'novembro'
        

    if mes2 == 12:
        mes3 = 'dezembro'
        
    return  x.strftime("%d") + " de " + mes3 + " de " + x.strftime("%Y")


def upload_spdoc(processo, txt):

    driver.switch_to.window(new_window)
    WebDriverWait(driver, 20).until(EC.title_contains("SP"))

    #===escolha do processo na mesa
    driver.implicitly_wait(2)
    driver.find_element_by_link_text(processo + "-A").click()


    #===Uploads de documentos
    driver.implicitly_wait(1)
    driver.find_element(By.CSS_SELECTOR, '[accesskey="d"]').click()

    driver.implicitly_wait(5)
    driver.find_element(By.ID, "dropdownMenuButton").click()

    driver.implicitly_wait(1)
    driver.find_element(By.CSS_SELECTOR, '[placeholder="Pesquisar modelo..."]').send_keys("cap" + Keys.DOWN + Keys.ENTER)

    processo = processo.replace("/", "-")

    driver.implicitly_wait(7)
    driver.find_element(By.ID, "Assunto").send_keys("Nota Fiscal Paulista")
    driver.implicitly_wait(2)
    driver.find_element(By.ID, "especie").send_keys("Extrato")
    driver.implicitly_wait(4)
    driver.find_element(By.ID, "conferencia").send_keys("Cópia autenticada administrativamente")
    driver.implicitly_wait(1)
    driver.find_element(By.ID, "arquivo").send_keys("C:\\Users\\Cliente\\OneDrive\\Documentos\\SEFAZ\\NFP\\NFP\\" + processo + "\\arquivo_completo.pdf")

    driver.implicitly_wait(1)
    driver.find_element(By.ID, "btnGravar").click()

    driver.implicitly_wait(12)
    driver.find_element(By.CSS_SELECTOR, '#page > div.row.siga-menu-acoes > div > p > a:nth-child(1)').click()

    driver.implicitly_wait(2)
    driver.find_element(By.ID, "bot-autenticar").click()

    driver.implicitly_wait(7)
    driver.find_element(By.ID, "senhaUsuarioSubscritor").send_keys(senha)

    driver.implicitly_wait(7)
    #driver.find_element(By.ID, "senhaOk").click()

    #WebDriverWait(driver, 30).until(lambda v: v.find_element_by_name('voltar'))
    #driver.find_element_by_name('voltar').click()
    sheet_ranges2.cell(linha, 12).value = "UPLOAD FEITO"
    wb.save('C:\\Users\\Cliente\\OneDrive\\Documentos\\SEFAZ\\NFP\\Gerador_de_Oficio2.xlsm')

    despacho_spdoc(processo, txt)

   


def despacho_spdoc(processo, txt):

    if driver.current_window_handle != new_window:
        driver.switch_to.window(new_window)
        (driver, 20).until(EC.title_contains("SP"))

        #===escolha do processo na mesa
        driver.implicitly_wait(2)
        driver.find_element_by_link_text(processo + "-A").click()



    #página principal do processo 
    driver.implicitly_wait(2)

    try:
        WebDriverWait(driver, 60).until(lambda t: t.find_element(By.CSS_SELECTOR, "#node2 a"))
        driver.find_element(By.CSS_SELECTOR, "#node2 a").click()
    except(e):
        driver.find_element(By.NAME, "voltar").click()
        driver.implicitly_wait(2)
        driver.find_element_by_link_text(processo + "-A").click()


    driver.implicitly_wait(5)
    driver.find_element(By.CSS_SELECTOR, '[accesskey="d"]').click()

    driver.implicitly_wait(2)
    driver.find_element(By.ID, "Assunto").send_keys("Nota Fiscal Paulista")
    #texto do despacho 
    driver.implicitly_wait(3)
    driver.execute_script('var ifram = document.getElementsByClassName("cke_wysiwyg_frame")[0];ifram.contentDocument.body.innerHTML = "' + txt + '";')

    driver.implicitly_wait(1)
    driver.find_element(By.ID, "btnGravar").click()

    driver.implicitly_wait(12)
    driver.find_element(By.CSS_SELECTOR, '#page > div.row.siga-menu-acoes > div > p > a:nth-child(1)').click()
 
    driver.implicitly_wait(2)
    driver.find_element(By.ID, "bot-assinar").click()

    driver.implicitly_wait(7)
    driver.find_element(By.ID, "senhaUsuarioSubscritor").send_keys(senha)

    driver.implicitly_wait(7)
    #driver.find_element(By.ID, "senhaOk").click()

    sheet_ranges2.cell(linha, 12).value = "CONCLUÍDO"
    wb.save("C:\\Users\\Cliente\\OneDrive\\Documentos\\SEFAZ\\NFP\\Gerador_de_Oficio2.xlsm")

    driver.find_element(By.NAME, "voltar").click()

    driver.switch_to.window(windows_before)
    WebDriverWait(driver, 20).until(EC.title_contains("F"))




#iniciador

qtd  = conta_linha('LINHA DE PRODUÇÃO')
for linha in range(2, qtd + 1):
    

    sheet_ranges2 = wb['LINHA DE PRODUÇÃO']
    Sigadoc = sheet_ranges2.cell(linha, 1).value.strip()
    Pedido = sheet_ranges2.cell(linha, 2).value.strip()
    Processo = sheet_ranges2.cell(linha, 3).value.strip()
    Data_do_oficio = sheet_ranges2.cell(linha, 4).value
    Oficio = sheet_ranges2.cell(linha, 7).value.strip()
    TituloOficio = sheet_ranges2.cell(linha, 9).value.strip()
    Juiz = sheet_ranges2.cell(linha, 10).value.strip()
    Vara = sheet_ranges2.cell(linha, 11).value.strip()
    Status = sheet_ranges2.cell(linha, 12).value.strip()

    Sigadoc_cru = Sigadoc #para ser usado no upload do processo no spdoc
    Sigadoc = Sigadoc.replace("/", "-")
    dados = lista_de_pesquisados(Sigadoc, Pedido)

    # SE O UPLOAD JÁ FOI FOI, PASSA-SE A DIANTE
    if Status == "UPLOAD FEITO" or Status == "CONCLUÍDO":

        #Se o uploado já foi feito o código vai direto para o despacho.
        if Status == "UPLOAD FEITO":
            if Pedido == "Transferência":
                #transferência mista
                if dados.find("Informamos que TODAS a(s) conta(s) já foi(foram) bloqueada(s). No presente momento, por imposição legal, a solicitação de transferência de valores feita por VSª poderá ser atendida SOMENTE COM RELAÇÃO AOS NOMES ACIMA, uma vez que há saldo(s) em conta NFP que é(são) inferior(es) a R$ 25,00 ou há pendências perante o CADIN estadual, relativo(s) a(os) seguinte(s) nome(s):") != -1:
                    final = dados.find("</table><br>")
                    dados_san = dados[0:final]

                    j = open("C:\\Users\\Cliente\\OneDrive\\Documentos\\Automa-o-spdoc\\modelo_despacho_DOF.txt", "r", encoding="utf-8")
                    texto_desp = j.read()
                
                    texto_desp = texto_desp.replace("&vara&", Vara)
                    texto_desp = texto_desp.replace("&juiz&", Juiz)
                    texto_desp = texto_desp.replace("&datof&", str(Data_do_oficio.day) + "/" + str(Data_do_oficio.month) + "/" + str(Data_do_oficio.year))
                    texto_desp = texto_desp.replace("&proc&", Processo)
                    texto_desp = texto_desp.replace("listadecontriba2", dados_san)
                    despacho_spdoc(Sigadoc_cru, texto_desp)

                else:
                    #transferência pura
                    j = open("C:\\Users\\Cliente\\OneDrive\\Documentos\\Automa-o-spdoc\\modelo_despacho_DOF.txt", "r", encoding="utf-8")
                    texto_desp = j.read()
                    dados = dados.replace("@24@", "")

                    texto_desp = texto_desp.replace("&vara&", Vara)
                    texto_desp = texto_desp.replace("&juiz&", Juiz)
                    texto_desp = texto_desp.replace("&datof&", str(Data_do_oficio.day) + "/" + str(Data_do_oficio.month) + "/" + str(Data_do_oficio.year))
                    texto_desp = texto_desp.replace("&proc&", Processo)
                    texto_desp = texto_desp.replace("listadecontriba2", dados)
                    despacho_spdoc(Sigadoc_cru, texto_desp)


            
            #Despacho de casos genéricos
            else:

                texto_desp = "<p>1. Expedido o E-Mail NFP DRT-5 nº fafnumfaf em resposta ao Ofício S/N de fafdataoffaf, referente ao Processo fafprocfaf.</p><p>2. Arquive-se.</p>"
                texto_desp = texto_desp.replace("fafnumfaf", Oficio)
                texto_desp = texto_desp.replace("fafdataoffaf", Data_do_oficio)
                texto_desp = texto_desp.replace("fafprocfaf", Processo)
                despacho_spdoc(Sigadoc_cru, texto_desp)


        
    else:
        #validador do loop - caso não haja somente transferências a serem encaminhadas para o DOF
        if dados.find("@24@") == -1:

            # Se já foi mandado e-mmail, não será enviado um duplicado
            if Status != "E-MAIL ENVIADO":
        
                #extrai os dados 

                
                dado_juiz = endereco_juiz(Juiz, Vara)
                Titulo1 = "ASSUNTO: RESPOSTA AO OFÍCIO S/N DE " + str(Data_do_oficio.day) + "/" + str(Data_do_oficio.month) + "/" + str(Data_do_oficio.year)
                Titulo2 = "PROCESSO JUDICIAL Nº " + Processo + " - SigaDOC "  +  Sigadoc

                #Manda o email
                f = open("C:\\Users\\Cliente\\OneDrive\\Documentos\\Automa-o-spdoc\\modelo_consulta.txt", "r", encoding="utf-8")
                texto = f.read()

                texto = texto.replace("juiznum", dado_juiz[0])
                texto = texto.replace("varanum", dado_juiz[1])
                texto = texto.replace("enderecovara", dado_juiz[2])
                texto = texto.replace("cepnum", dado_juiz[3])
                texto = texto.replace("ofnumfaf", Oficio)
                texto = texto.replace("datadehj", mes())
                texto = texto.replace("titulo1", Titulo1)
                texto = texto.replace("titulo2", Titulo2)
                texto = texto.replace("listadecontriba", dados)
                bloqueiotransferencia = str()
                if Pedido == 'Transferência':
                    if texto.find("@25@") != -1:
                        texto.replace("@25@", "")
                        bloqueiotransferencia = "Informamos que a(s) conta(s) já foi(foram) bloqueada(s). No presente momento, por imposição legal, a solicitação de transferência de valores feita por VSª não poderá ser atendida, uma vez que há saldo(s) em conta NFP que é(são) inferior(es) a R$ 25,00 e(ou) o há pendências no CADIN."
                    elif texto.find("@24@") != -1:
                        texto.replace("@24@", "")
                        bloqueiotransferencia = "NÃO ENCAMINHAR ESSE E-MAIL."
                    else:
                        bloqueiotransferencia = ""
                
                elif Pedido == 'Bloqueio':
                    bloqueiotransferencia = 'Adicionalmente, informamos que a(s) conta(s) já foi(foram) bloqueada(s).'
                else:
                    bloqueiotransferencia = ''

                texto = texto.replace("bloqueiotransferencia", bloqueiotransferencia)

                

                #driver.implicitly_wait(3)
                #driver.find_element(By.CSS_SELECTOR, '[autoid="_n_j"]').click()

                driver.implicitly_wait(5)
                #email
                WebDriverWait(driver, timeout=20).until(lambda d: d.find_element(By.CSS_SELECTOR, '[autoid="_fp_7"]'))
                driver.find_element(By.CSS_SELECTOR, '[autoid="_fp_7"]').send_keys(dado_juiz[4])
            
                driver.implicitly_wait(5)
                driver.find_element(By.CSS_SELECTOR, '[aria-label="Destinatários Cc. Digite um endereço de email ou um nome de sua lista de contatos."]').send_keys('delegado05@fazenda.sp.gov.br')

                #assunto
                driver.implicitly_wait(5)
                driver.find_element(By.CSS_SELECTOR, '[autoid="_mcp_k"]').send_keys(Titulo1 + " " + Titulo2)

            
                #corpo do texto
                driver.implicitly_wait(7)
                driver.find_element(By.CSS_SELECTOR, '#EditorBody').click()
                WebDriverWait(driver, timeout=30).until(lambda e: e.find_element(By.CSS_SELECTOR, '#EditorBody'))
                driver.execute_script('var iframe = document.getElementById("EditorBody");var quadro = iframe.contentWindow.document.getElementById("MicrosoftOWAEditorRegion");var paragrafos = quadro.getElementsByTagName("p"); paragrafos[0].innerHTML = "' + texto + '";')


                #enviar
                driver.implicitly_wait(5)
                #driver.find_element(By.CSS_SELECTOR, '[autoid="_mcp_6"]').click()
                

                driver.find_element(By.CSS_SELECTOR, '[autoid="_n_j"]').click()

                sheet_ranges2.cell(linha, 12).value = "E-MAIL ENVIADO"
                wb.save('C:\\Users\\Cliente\\OneDrive\\Documentos\\SEFAZ\\NFP\\Gerador_de_Oficio2.xlsm')


            #Será  feito o despacho e o upload
            #despacho de transferência mista
            if dados.find("Informamos que TODAS a(s) conta(s) já foi(foram) bloqueada(s). No presente momento, por imposição legal, a solicitação de transferência de valores feita por VSª poderá ser atendida SOMENTE COM RELAÇÃO AOS NOMES ACIMA, uma vez que há saldo(s) em conta NFP que é(são) inferior(es) a R$ 25,00 ou há pendências perante o CADIN estadual, relativo(s) a(os) seguinte(s) nome(s):") != -1:
                final = dados.find("</table><br>")
                dados_san = dados[0:final]

                j = open("C:\\Users\\Cliente\\OneDrive\\Documentos\\Automa-o-spdoc\\modelo_despacho_DOF.txt", "r", encoding="utf-8")
                texto_desp = j.read()
            
                texto_desp = texto_desp.replace("&vara&", Vara)
                texto_desp = texto_desp.replace("&juiz&", Juiz)
                texto_desp = texto_desp.replace("&datof&", str(Data_do_oficio.day) + "/" + str(Data_do_oficio.month) + "/" + str(Data_do_oficio.year))
                texto_desp = texto_desp.replace("&proc&", Processo)
                texto_desp = texto_desp.replace("listadecontriba2", dados_san)
                upload_spdoc(Sigadoc_cru, texto_desp)

            #demais casos
            else:
                texto_desp = "<p>1. Expedido o E-Mail NFP DRT-5 nº fafnumfaf em resposta ao Ofício S/N de fafdataoffaf, referente ao Processo fafprocfaf.</p><p>2. Arquive-se.</p>"
                texto_desp = texto_desp.replace("fafnumfaf", Oficio)
                texto_desp = texto_desp.replace("fafdataoffaf", str(Data_do_oficio.day) + "/" + str(Data_do_oficio.month) + "/" + str(Data_do_oficio.year))
                texto_desp = texto_desp.replace("fafprocfaf", Processo)
                upload_spdoc(Sigadoc_cru, texto_desp)


            reply = str(input('Deu tudo  certo? (y/n):')).lower().strip()
            if reply[0] == 'y':

                pass
            elif reply[0] == 'n':
                sys.exit()

            else:
                print("prosseguindo...")

        #caso seja transferência pura
        else:
            
            # &proc& &datof& &juiz& &vara& listadecontriba2
            
            j = open("C:\\Users\\Cliente\\OneDrive\\Documentos\\Automa-o-spdoc\\modelo_despacho_DOF.txt", "r", encoding="utf-8")
            texto_desp = j.read()
            dados = dados.replace("@24@", "")

            texto_desp = texto_desp.replace("&vara&", Vara)
            texto_desp = texto_desp.replace("&juiz&", Juiz)
            texto_desp = texto_desp.replace("&datof&", str(Data_do_oficio.day) + "/" + str(Data_do_oficio.month) + "/" + str(Data_do_oficio.year))
            texto_desp = texto_desp.replace("&proc&", Processo)
            texto_desp = texto_desp.replace("listadecontriba2", dados)
            upload_spdoc(Sigadoc_cru, texto_desp)

            reply = str(input('Deu tudo  certo? (y/n):')).lower().strip()
            if reply[0] == 'y':

                pass
            elif reply[0] == 'n':
                sys.exit()

            else:
                print("prosseguindo...")
    


